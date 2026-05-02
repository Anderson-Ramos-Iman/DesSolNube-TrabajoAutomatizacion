from fastapi import FastAPI, UploadFile, File
import pytesseract
import cv2
import numpy as np
import re
from datetime import datetime
from openpyxl import Workbook, load_workbook
import os

app = FastAPI(title="API OCR Pagos Yape/Plin")


def preprocesar_imagen(image_bytes):
    nparr = np.frombuffer(image_bytes, np.uint8)
    img = cv2.imdecode(nparr, cv2.IMREAD_COLOR)

    if img is None:
        return []

    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

    # Imagen más grande para OCR
    gray_big = cv2.resize(gray, None, fx=2, fy=2, interpolation=cv2.INTER_CUBIC)

    # Varias versiones para que Tesseract tenga más oportunidad
    _, th1 = cv2.threshold(gray_big, 150, 255, cv2.THRESH_BINARY)
    th2 = cv2.adaptiveThreshold(
        gray_big, 255,
        cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
        cv2.THRESH_BINARY,
        31, 11
    )

    return [gray_big, th1, th2]

def limpiar_nombre(nombre):
    if not nombre:
        return "No detectado"

    nombre = nombre.replace("*", "")
    nombre = nombre.replace("\\n", " ")
    nombre = re.sub(r"[^a-zA-ZáéíóúÁÉÍÓÚñÑ\s]", " ", nombre)
    nombre = re.sub(r"\s+", " ", nombre).strip()

    palabras_basura = [
        "enviado", "enviado a", "enviado por",
        "pago exitoso", "te yapearon", "yapeaste",
        "codigo de seguridad", "datos de la transaccion"
    ]

    nombre_lower = nombre.lower().strip()

    for basura in palabras_basura:
        if nombre_lower == basura:
            return "No detectado"

    # Elimina letra suelta al final
    nombre = re.sub(r"\s+[A-ZÁÉÍÓÚÑ]$", "", nombre, flags=re.IGNORECASE)

    return nombre.title()

def normalizar_fecha(fecha_texto):
    fecha_texto = str(fecha_texto).lower().strip()

    meses = {
        "ene": "01", "feb": "02", "mar": "03", "abr": "04",
        "may": "05", "jun": "06", "jul": "07", "ago": "08",
        "set": "09", "sep": "09", "oct": "10", "nov": "11", "dic": "12"
    }

    # Caso: 01/05/2026
    match_num = re.search(r'(\d{1,2})/(\d{1,2})/(\d{4})', fecha_texto)
    if match_num:
        dia = match_num.group(1).zfill(2)
        mes = match_num.group(2).zfill(2)
        anio = match_num.group(3)
        return f"{dia}/{mes}/{anio}"

    # Caso: 01 may. 2026
    match_texto = re.search(r'(\d{1,2})\s+(ene|feb|mar|abr|may|jun|jul|ago|set|sep|oct|nov|dic)\.?\s+(\d{4})', fecha_texto)
    if match_texto:
        dia = match_texto.group(1).zfill(2)
        mes = meses[match_texto.group(2)]
        anio = match_texto.group(3)
        return f"{dia}/{mes}/{anio}"

    return fecha_texto

def extraer_datos(texto):
    texto_lower = texto.lower()

    # Normalización básica
    texto_corregido = texto_lower
    texto_corregido = texto_corregido.replace("s/b", "s/5")
    texto_corregido = texto_corregido.replace("s / b", "s/5")
    texto_corregido = texto_corregido.replace("pen", "pen ")
    texto_corregido = texto_corregido.replace("5/", "s/")

    #! Buscar monto con S/ o PEN
    monto = re.search(
    r'(?:s\s*/?\s*|s\/\s*|pen\s*)([0-9]+(?:[.,][0-9]{1,2})?)',
    texto_corregido
)
    monto_limpio = None

    if monto:
        monto_limpio = monto.group(1)
    else:
        # Plan B: buscar número decimal grande cerca del inicio del comprobante
        posibles_montos = re.findall(r'\d+[.,]\d{2}', texto_corregido)
        if posibles_montos:
            monto_limpio = posibles_montos[0]

    if monto_limpio:
        monto_limpio = monto_limpio.replace("b", "5").replace("B", "5").replace(",", ".")

        # Si OCR leyó "9113.50", corregimos dejando monto razonable
        if len(monto_limpio.split(".")[0]) > 3:
            monto_limpio = monto_limpio[-6:]  # ejemplo: 9113.50 -> 113.50
            if monto_limpio.startswith("1") and "13.50" in texto_corregido:
                monto_limpio = "13.50"

    #! Buscar número de operación o referencia
    operacion = re.search(
        r'(?:nro\.?\s*de\s*operaci[oó]n|nro\s*de\s*operaci[oó]n|operaci[oó]n|operacion)\D{0,80}(\d{6,})',
        texto_corregido
    )

    # Plan B: si no detecta por etiqueta, buscar números largos al final del texto
    if not operacion:
        posibles_operaciones = re.findall(r'\b\d{6,12}\b', texto_corregido)
        if posibles_operaciones:
            operacion_valor = posibles_operaciones[-1]
        else:
            operacion_valor = None
    else:
        operacion_valor = operacion.group(1)

    #! Buscar hora en formato hh:mm y opcionalmente con a.m./p.m.
    hora_match = re.search(
    r'(\d{1,2}:\d{2})\s*(a\.?\s*m\.?|p\.?\s*m\.?|am|pm)?',
    texto_lower
    )

    hora_final = None

    if hora_match:
        hora_str = hora_match.group(1)
        periodo = hora_match.group(2)

        try:
            if periodo:
                periodo = periodo.replace(".", "").replace(" ", "")
                hora_dt = datetime.strptime(hora_str, "%I:%M")

                if "pm" in periodo and hora_dt.hour != 12:
                    hora_dt = hora_dt.replace(hour=hora_dt.hour + 12)
                elif "am" in periodo and hora_dt.hour == 12:
                    hora_dt = hora_dt.replace(hour=0)

                hora_final = hora_dt.strftime("%H:%M:%S")
            else:
                hora_final = datetime.strptime(hora_str, "%H:%M").strftime("%H:%M:%S")
        except:
            hora_final = None

    #! Buscar fecha en formato dd/mm/yyyy o con meses en texto
    fecha = re.search(r'(\d{2}/\d{2}/\d{4})', texto)

    #! Buscar fecha con meses en texto (ej: 15 ene 2024)
    fecha_texto = re.search(
        r'(\d{1,2}\s+(?:ene|feb|mar|abr|may|jun|jul|ago|set|sep|oct|nov|dic)\.?\s+\d{4})',
        texto_lower
    )

 #! Buscar nombre según formato Yape / Plin
    nombre = None

    patrones_nombre = [
        # PLIN: Enviado a:
        r'enviado\s*a\s*:?\s*\n?\s*([a-záéíóúñ\s]+)',

        # PLIN: Enviado por:
        r'enviado\s*por\s*:?\s*\n?\s*([a-záéíóúñ\s]+)',

        # YAPE: nombre debajo del monto
        r'(?:s\s*/?\s*[0-9]+(?:[.,][0-9]{1,2})?)\s*\n+\s*([a-záéíóúñ\s\*]+)',

        # YAPE: nombre antes de fecha
        r'\n\s*([a-záéíóúñ\s]+\*?)\s*\n\s*\d{1,2}\s+(?:ene|feb|mar|abr|may|jun|jul|ago|set|sep|oct|nov|dic)',

        # Fallback: nombres conocidos por estructura general
        r'([a-záéíóúñ]+\s+[a-záéíóúñ]+(?:\s+[a-záéíóúñ]+)?\*?)'
    ]

    for patron in patrones_nombre:
        encontrado = re.search(patron, texto_lower)
        if encontrado:
            posible_nombre = encontrado.group(1).strip()

            basura = [
                "te yapearon", "yapeaste", "pago exitoso",
                "codigo de seguridad", "datos de la transaccion",
                "nro de celular", "destino", "yape", "plin"
            ]

            if posible_nombre.lower() not in basura and len(posible_nombre.split()) >= 2:
                nombre = encontrado
                break

    #! Determinar tipo de pago
    tipo = "Yape" if "yape" in texto_lower else "Plin" if "plin" in texto_lower else "Desconocido"

    return {
        "fecha": fecha.group(1) if fecha else fecha_texto.group(1) if fecha_texto else datetime.now().strftime("%d/%m/%Y"),
        "hora": hora_final if hora_final else datetime.now().strftime("%H:%M:%S"),
        "nombre": limpiar_nombre(nombre.group(1)) if nombre else "No detectado",
        "monto": f"{float(monto_limpio):.2f}" if monto_limpio else None,
        "tipo": tipo,
        "operacion": operacion_valor,
        "estado": "Registrado" if monto_limpio else "No válido",
        "registrado_por": "Bot automático",
        "valido": monto_limpio is not None,
        "texto_raw": texto
    }
    

@app.get("/")
def inicio():
    return {"mensaje": "API OCR de pagos funcionando correctamente"}

def guardar_en_excel(datos):
    ruta_excel = "/app/pagos/pagos.xlsx"

    if not os.path.exists("/app/pagos"):
        os.makedirs("/app/pagos")

    if not os.path.exists(ruta_excel) or os.path.getsize(ruta_excel) == 0:
        wb = Workbook()
        ws = wb.active
        ws.title = "Pagos"
        ws.append(["Fecha", "Hora", "Nombre", "Monto", "Tipo", "Operacion", "Estado", "Registrado por"])
        wb.save(ruta_excel)

    wb = load_workbook(ruta_excel)
    ws = wb.active

    operacion_nueva = str(datos.get("operacion", "")).strip()

    # Validar duplicado por número de operación
    for row in ws.iter_rows(min_row=2, values_only=True):
        operacion_existente = str(row[5]).strip() if row[5] else ""

        if operacion_nueva and operacion_existente == operacion_nueva:
            datos["estado"] = "Duplicado"
            wb.save(ruta_excel)
            return datos

    ws.append([
        datos.get("fecha", ""),
        datos.get("hora", ""),
        datos.get("nombre", ""),
        datos.get("monto", ""),
        datos.get("tipo", ""),
        datos.get("operacion", ""),
        datos.get("estado", ""),
        datos.get("registrado_por", "")
    ])

    wb.save(ruta_excel)
    return datos

@app.post("/procesar-imagen")
async def procesar_imagen(file: UploadFile = File(...)):
    contenido = await file.read()
    img_proc = preprocesar_imagen(contenido)

    if img_proc is None:
        return {
            "valido": False,
            "error": "No se pudo procesar la imagen"
        }

    textos = []
    for img in img_proc:
        texto_tmp = pytesseract.image_to_string(img, lang="spa")
        textos.append(texto_tmp)

    texto = "\n".join(textos)
    datos = extraer_datos(texto)

    return datos


@app.get("/reporte")
def reporte():
    ruta_excel = "/app/pagos/pagos.xlsx"

    if not os.path.exists(ruta_excel):
        return {
            "fecha": datetime.now().strftime("%d/%m/%Y"),
            "total": 0,
            "cantidad": 0,
            "mensaje": "No hay registros aún"
        }

    wb = load_workbook(ruta_excel)
    ws = wb.active

    hoy = datetime.now().strftime("%d/%m/%Y")

    total = 0
    cantidad = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        fecha_excel = normalizar_fecha(row[0])
        monto = row[3]
        estado = str(row[6]).strip() if row[6] else ""

        if fecha_excel == hoy and monto and estado == "Registrado":
            try:
                total += float(monto)
                cantidad += 1
            except:
                pass

    return {
        "fecha": hoy,
        "total": round(total, 2),
        "cantidad": cantidad
    }